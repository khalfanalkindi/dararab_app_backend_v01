#!/usr/bin/env python3
"""
Minimal contract import (one-off).

Imports only requested fields from spreadsheet and leaves most FKs/manual fields for later:
- yes: title, project_title_ar(resolve to project_id), commission_percent, fixed_amount,
       free_copies, contract_duration, start_date, end_date, payment_schedule, notes
- no:  contract_type/status/royalties/signed_by (left NULL)

Important:
Contract model requires GenericFK backing fields (content_type_id, object_id), so this script
uses a configurable placeholder party (default: rightsowner id=2) until users adjust manually.

Usage:
  python scripts/import_contracts_minimal_mysql.py --dry-run
  python scripts/import_contracts_minimal_mysql.py
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence
from urllib.parse import unquote, urlparse

try:
    import mysql.connector
except ImportError as e:
    print("Install: pip install mysql-connector-python", file=sys.stderr)
    raise e


MYSQL_URL = (
    "mysql://root:DPMCwYKeEPBvcGdQheifhQqICRUuGzBL@mainline.proxy.rlwy.net:29174/railway"
)
MYSQL_HOST = "mainline.proxy.rlwy.net"
MYSQL_PORT = 29174
MYSQL_USER = "root"
MYSQL_PASSWORD = "DPMCwYKeEPBvcGdQheifhQqICRUuGzBL"
MYSQL_DATABASE = "railway"
CREATED_BY_ID = 4
UPDATED_BY_ID = None

# Required for inventory_contract generic relation (manual cleanup can happen later).
DEFAULT_CONTRACTED_PARTY_TYPE = "rightsowner"  # author|translator|rightsowner|reviewer|stakeholder
DEFAULT_CONTRACTED_PARTY_ID = 2

_SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_FILE = _SCRIPT_DIR / "contracts_import.numbers"

HEADER_REQUIRED = {"title", "project_title_ar"}
HEADER_ALL = [
    "title",
    "project_title_ar",
    "commission_percent",
    "fixed_amount",
    "free_copies",
    "contract_duration",
    "start_date",
    "end_date",
    "payment_schedule",
    "notes",
]


def _parse_mysql_url(url: str) -> dict[str, Any]:
    url = (url or "").strip()
    if not url:
        return {}
    if url.startswith("mysql+pymysql://"):
        url = "mysql://" + url[len("mysql+pymysql://") :]
    elif url.startswith("mysql2://"):
        url = "mysql://" + url[len("mysql2://") :]
    elif not url.startswith("mysql://"):
        return {}
    u = urlparse(url)
    if not u.hostname:
        return {}
    db = (u.path or "").lstrip("/").split("?", 1)[0]
    return {
        "host": u.hostname,
        "port": int(u.port or 3306),
        "user": unquote(u.username or ""),
        "password": unquote(u.password or ""),
        "database": db,
    }


def _connect_kwargs() -> dict[str, Any]:
    raw = os.environ.get("MYSQL_URL") or os.environ.get("DATABASE_URL") or MYSQL_URL
    p = _parse_mysql_url(str(raw))
    return {
        "host": os.environ.get("MYSQL_HOST") or p.get("host") or MYSQL_HOST,
        "port": int(os.environ.get("MYSQL_PORT") or p.get("port") or MYSQL_PORT),
        "user": os.environ.get("MYSQL_USER") or p.get("user") or MYSQL_USER,
        "password": os.environ.get("MYSQL_PASSWORD") or p.get("password") or MYSQL_PASSWORD,
        "database": os.environ.get("MYSQL_DATABASE") or p.get("database") or MYSQL_DATABASE,
        "charset": "utf8mb4",
        "use_unicode": True,
    }


def _resolve_input_path(arg_file: Optional[str]) -> Path:
    env_file = os.environ.get("CONTRACT_IMPORT_FILE")
    if arg_file:
        return Path(arg_file).expanduser().resolve()
    if env_file:
        return Path(env_file).expanduser().resolve()
    if DEFAULT_INPUT_FILE.is_file():
        return DEFAULT_INPUT_FILE.resolve()
    raise SystemExit(
        f"No contracts input file found at {DEFAULT_INPUT_FILE}. "
        "Use --file or set CONTRACT_IMPORT_FILE."
    )


def _norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = " ".join(str(v).split()).strip()
    return s or None


def _to_decimal(v: Any) -> Optional[float]:
    s = _norm(v)
    if s is None:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _to_int(v: Any) -> Optional[int]:
    s = _norm(v)
    if s is None:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _to_date(v: Any) -> Optional[str]:
    s = _norm(v)
    if not s:
        return None
    fmts = ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d")
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _matrix_to_rows(header_row: Sequence[Any], body_rows: Iterable[Sequence[Any]]) -> list[dict[str, Any]]:
    headers = ["" if h is None else str(h).strip().lower() for h in header_row]
    missing = HEADER_REQUIRED - set(headers)
    if missing:
        raise ValueError(f"Missing required headers: {sorted(missing)}. Found: {headers}")
    idx = {h: i for i, h in enumerate(headers) if h}
    out: list[dict[str, Any]] = []
    for r in body_rows:
        if r is None:
            continue
        row = list(r)
        if not row or all(_norm(x) is None for x in row):
            continue
        rec = {}
        for k in HEADER_ALL:
            i = idx.get(k)
            rec[k] = row[i] if i is not None and i < len(row) else None
        if _norm(rec.get("title")) is None and _norm(rec.get("project_title_ar")) is None:
            continue
        out.append(rec)
    return out


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("CSV has no header row")
            headers = [h for h in reader.fieldnames if h]
            matrix = [[row.get(h) for h in headers] for row in reader]
            return _matrix_to_rows(headers, matrix)
    if suf in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise SystemExit("For .xlsx install: pip install openpyxl") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            wb.close()
            raise ValueError("Empty workbook")
        data = _matrix_to_rows(header, list(rows))
        wb.close()
        return data
    if suf == ".numbers":
        try:
            from numbers_parser import Document
        except ImportError as e:
            raise SystemExit("For .numbers install: pip install numbers-parser") from e
        doc = Document(str(path))
        if not doc.sheets or not doc.sheets[0].tables:
            raise ValueError("Numbers file needs at least one sheet/table")
        matrix = doc.sheets[0].tables[0].rows(values_only=True)
        if not matrix:
            raise ValueError("Numbers table is empty")
        return _matrix_to_rows(matrix[0], matrix[1:])
    raise ValueError(f"Unsupported input file type: {path.suffix!r}")


def main() -> int:
    p = argparse.ArgumentParser(description="Import minimal contracts and print inserted/skipped stats.")
    p.add_argument("--dry-run", action="store_true", help="Validate only and rollback.")
    p.add_argument("--file", "-f", default=None, help="Path to .numbers/.csv/.xlsx file.")
    args = p.parse_args()

    path = _resolve_input_path(args.file)
    rows = _read_rows(path)
    if not rows:
        print("No rows found.")
        return 0

    created_by = int(os.environ.get("CREATED_BY_ID", str(CREATED_BY_ID)))
    updated_by_raw = os.environ.get("UPDATED_BY_ID", "")
    updated_by = int(updated_by_raw) if updated_by_raw.strip() else UPDATED_BY_ID
    party_type = os.environ.get("DEFAULT_CONTRACTED_PARTY_TYPE", DEFAULT_CONTRACTED_PARTY_TYPE)
    party_id = int(os.environ.get("DEFAULT_CONTRACTED_PARTY_ID", str(DEFAULT_CONTRACTED_PARTY_ID)))

    cnx = mysql.connector.connect(autocommit=False, **_connect_kwargs())
    cur = cnx.cursor()

    inserted = 0
    skipped = 0
    reasons: dict[str, int] = {}
    skipped_rows: list[dict[str, Any]] = []

    def skip(reason: str, row_no: int, title: str, project_title: Optional[str]) -> None:
        nonlocal skipped
        skipped += 1
        reasons[reason] = reasons.get(reason, 0) + 1
        skipped_rows.append(
            {
                "row_no": row_no,
                "reason": reason,
                "title": title or "",
                "project_title_ar": project_title or "",
            }
        )

    try:
        # resolve content type for required GFK columns
        cur.execute(
            "SELECT id FROM django_content_type WHERE app_label='inventory' AND model=%s LIMIT 1",
            (party_type,),
        )
        row = cur.fetchone()
        if not row:
            raise ValueError(f"No django_content_type for inventory.{party_type}")
        content_type_id = int(row[0])

        insert_sql = """
            INSERT INTO inventory_contract (
              created_at, updated_at,
              title, project_id, content_type_id, object_id,
              commission_percent, fixed_amount, free_copies,
              contract_duration, start_date, end_date,
              payment_schedule, notes,
              created_by_id, updated_by_id
            ) VALUES (
              SYSDATE(6), SYSDATE(6),
              %s, %s, %s, %s,
              %s, %s, %s,
              %s, %s, %s,
              %s, %s,
              %s, %s
            )
        """

        for i, rec in enumerate(rows, start=2):
            title = _norm(rec.get("title")) or ""
            project_title = _norm(rec.get("project_title_ar"))
            if not project_title:
                skip("missing_project_title_ar", i, title, project_title)
                continue

            cur.execute(
                "SELECT id FROM inventory_project WHERE LOWER(TRIM(title_ar)) = LOWER(TRIM(%s)) LIMIT 1",
                (project_title,),
            )
            pr = cur.fetchone()
            if not pr:
                skip("project_not_found", i, title, project_title)
                continue
            project_id = int(pr[0])

            commission = _to_decimal(rec.get("commission_percent"))
            fixed = _to_decimal(rec.get("fixed_amount"))
            free_copies = _to_int(rec.get("free_copies"))
            duration = _to_int(rec.get("contract_duration"))
            start_date = _to_date(rec.get("start_date"))
            end_date = _to_date(rec.get("end_date"))
            payment_schedule = _norm(rec.get("payment_schedule")) or ""
            notes = _norm(rec.get("notes"))

            if _norm(rec.get("start_date")) and start_date is None:
                skip("invalid_start_date", i, title, project_title)
                continue
            if _norm(rec.get("end_date")) and end_date is None:
                skip("invalid_end_date", i, title, project_title)
                continue

            cur.execute(
                insert_sql,
                (
                    title,
                    project_id,
                    content_type_id,
                    party_id,
                    commission,
                    fixed,
                    free_copies,
                    duration,
                    start_date,
                    end_date,
                    payment_schedule,
                    notes,
                    created_by,
                    updated_by,
                ),
            )
            inserted += 1

        if args.dry_run:
            cnx.rollback()
        else:
            cnx.commit()

        print(f"{'Dry-run' if args.dry_run else 'Done'}: total={len(rows)}, inserted={inserted}, skipped={skipped}")
        if reasons:
            print("Skipped reasons:")
            for k, v in sorted(reasons.items(), key=lambda x: (-x[1], x[0])):
                print(f"  - {k}: {v}")
        if skipped_rows:
            print("\nSkipped rows details:")
            for r in skipped_rows:
                print(
                    f"  - row~{r['row_no']}: reason={r['reason']}, "
                    f"title={r['title']!r}, project_title_ar={r['project_title_ar']!r}"
                )

    except Exception as e:
        cnx.rollback()
        print(f"Rolled back. Error: {e}", file=sys.stderr)
        return 1
    finally:
        cur.close()
        cnx.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())

