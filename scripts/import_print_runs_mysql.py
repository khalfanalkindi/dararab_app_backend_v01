#!/usr/bin/env python3
"""
Import print runs from Numbers/CSV/XLSX into inventory_printrun.

Resolves product by ISBN only (column product_title_ar is ignored).
Status (status_text): matches common_listitem under list_type code printrun_status by
  list item id, value, display_name_en, or display_name_ar (case-insensitive trim).
If status_text is empty, status_id is NULL.

Default input path is PRINT_RUN_INPUT_PATH below (same folder as this script).
Copy your .numbers file there, or edit that constant. Optional override: --file

Usage:
  python scripts/import_print_runs_mysql.py --dry-run
  python scripts/import_print_runs_mysql.py --file /path/to/other.numbers
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence
from urllib.parse import unquote, urlparse

try:
    import mysql.connector
except ImportError as e:
    print("Install: pip install mysql-connector-python", file=sys.stderr)
    raise e

MYSQL_URL = (
    "mysql://root:DPMCwYKeEPBvcGdQheifhQqICRUuGzBL@mainline.proxy.rlwy.net:29174/railway"
)
MYSQL_HOST = "mainline.proxy.rlwy.net"
MYSQL_PORT = 29174
MYSQL_USER = "root"
MYSQL_PASSWORD = "DPMCwYKeEPBvcGdQheifhQqICRUuGzBL"
MYSQL_DATABASE = "railway"
CREATED_BY_ID = 4
UPDATED_BY_ID: Optional[int] = None

LIST_TYPE_CODE_PRINT_RUN_STATUS = "printrun_status"

_SCRIPT_DIR = Path(__file__).resolve().parent
# Put your sheet in scripts/ with this name, or set an absolute path here.
PRINT_RUN_INPUT_PATH = _SCRIPT_DIR / "print_runs_import.numbers"

HEADER_REQUIRED = {"product_isbn"}
HEADER_ALL = [
    "product_isbn",
    "product_title_ar",
    "edition_number",
    "price",
    "price_omr",
    "status_text",
    "published_at",
    "notes",
]


def _parse_mysql_url(url: str) -> dict[str, Any]:
    url = (url or "").strip()
    if not url:
        return {}
    if url.startswith("mysql+pymysql://"):
        url = "mysql://" + url[len("mysql+pymysql://") :]
    elif url.startswith("mysql2://"):
        url = "mysql://" + url[len("mysql2://") :]
    elif not url.startswith("mysql://"):
        return {}
    u = urlparse(url)
    if not u.hostname:
        return {}
    db = (u.path or "").lstrip("/").split("?", 1)[0]
    return {
        "host": u.hostname,
        "port": int(u.port or 3306),
        "user": unquote(u.username or ""),
        "password": unquote(u.password or ""),
        "database": db,
    }


def _connect_kwargs() -> dict[str, Any]:
    raw = os.environ.get("MYSQL_URL") or os.environ.get("DATABASE_URL") or MYSQL_URL
    p = _parse_mysql_url(str(raw))
    return {
        "host": os.environ.get("MYSQL_HOST") or p.get("host") or MYSQL_HOST,
        "port": int(os.environ.get("MYSQL_PORT") or p.get("port") or MYSQL_PORT),
        "user": os.environ.get("MYSQL_USER") or p.get("user") or MYSQL_USER,
        "password": os.environ.get("MYSQL_PASSWORD") or p.get("password") or MYSQL_PASSWORD,
        "database": os.environ.get("MYSQL_DATABASE") or p.get("database") or MYSQL_DATABASE,
        "charset": "utf8mb4",
        "use_unicode": True,
    }


def _resolve_input_path(arg_file: Optional[str]) -> Path:
    if arg_file:
        return Path(arg_file).expanduser().resolve()
    if PRINT_RUN_INPUT_PATH.is_file():
        return PRINT_RUN_INPUT_PATH.resolve()
    raise SystemExit(
        f"No print run input file found at:\n  {PRINT_RUN_INPUT_PATH}\n"
        "Copy your .numbers file there, edit PRINT_RUN_INPUT_PATH in this script, "
        "or pass --file /path/to/your.numbers"
    )


def _norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = " ".join(str(v).split()).strip()
    return s or None


def _norm_isbn_key(v: Any) -> Optional[str]:
    s = _norm(v)
    if not s:
        return None
    return "".join(c for c in s.lower() if c.isalnum())


def _to_decimal(v: Any) -> Optional[float]:
    s = _norm(v)
    if s is None:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _to_int(v: Any) -> Optional[int]:
    s = _norm(v)
    if s is None:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _to_date(v: Any) -> Optional[str]:
    s = _norm(v)
    if not s:
        return None
    fmts = ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d")
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _matrix_to_rows(header_row: Sequence[Any], body_rows: Iterable[Sequence[Any]]) -> list[dict[str, Any]]:
    headers = ["" if h is None else str(h).strip().lower() for h in header_row]
    missing = HEADER_REQUIRED - set(headers)
    if missing:
        raise ValueError(f"Missing required headers: {sorted(missing)}. Found: {headers}")
    idx = {h: i for i, h in enumerate(headers) if h}
    out: list[dict[str, Any]] = []
    for r in body_rows:
        if r is None:
            continue
        row = list(r)
        if not row or all(_norm(x) is None for x in row):
            continue
        rec: dict[str, Any] = {}
        for k in HEADER_ALL:
            i = idx.get(k)
            rec[k] = row[i] if i is not None and i < len(row) else None
        if _norm_isbn_key(rec.get("product_isbn")) is None:
            continue
        out.append(rec)
    return out


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("CSV has no header row")
            headers = [h for h in reader.fieldnames if h]
            matrix = [[row.get(h) for h in headers] for row in reader]
            return _matrix_to_rows(headers, matrix)
    if suf in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise SystemExit("For .xlsx install: pip install openpyxl") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            wb.close()
            raise ValueError("Empty workbook")
        data = _matrix_to_rows(header, list(rows))
        wb.close()
        return data
    if suf == ".numbers":
        try:
            from numbers_parser import Document
        except ImportError as e:
            raise SystemExit("For .numbers install: pip install numbers-parser") from e
        doc = Document(str(path))
        if not doc.sheets or not doc.sheets[0].tables:
            raise ValueError("Numbers file needs at least one sheet/table")
        matrix = doc.sheets[0].tables[0].rows(values_only=True)
        if not matrix:
            raise ValueError("Numbers table is empty")
        return _matrix_to_rows(matrix[0], matrix[1:])
    raise ValueError(f"Unsupported input file type: {path.suffix!r}")


def _resolve_product_id(cur: Any, isbn_key: str) -> Optional[int]:
    cur.execute(
        """
        SELECT id FROM inventory_product
        WHERE REPLACE(REPLACE(LOWER(TRIM(isbn)), '-', ''), ' ', '') = %s
        LIMIT 1
        """,
        (isbn_key,),
    )
    row = cur.fetchone()
    return int(row[0]) if row else None


def _resolve_status_id(cur: Any, status_text: Optional[str]) -> tuple[Optional[int], Optional[str]]:
    """
    Returns (status_id, error_reason_if_any).
    Empty status -> (None, None).
    """
    raw = _norm(status_text)
    if raw is None:
        return None, None

    try:
        maybe_id = int(float(raw))
    except (TypeError, ValueError):
        maybe_id = None
    if maybe_id is not None:
        cur.execute(
            """
            SELECT li.id FROM common_listitem li
            INNER JOIN common_listtype lt ON lt.id = li.list_type_id
            WHERE lt.code = %s AND li.is_active = 1 AND li.id = %s
            LIMIT 1
            """,
            (LIST_TYPE_CODE_PRINT_RUN_STATUS, maybe_id),
        )
        row = cur.fetchone()
        if row:
            return int(row[0]), None

    cur.execute(
        """
        SELECT li.id FROM common_listitem li
        INNER JOIN common_listtype lt ON lt.id = li.list_type_id
        WHERE lt.code = %s AND li.is_active = 1
          AND (
            LOWER(TRIM(li.value)) = LOWER(TRIM(%s))
            OR LOWER(TRIM(li.display_name_en)) = LOWER(TRIM(%s))
            OR LOWER(TRIM(li.display_name_ar)) = LOWER(TRIM(%s))
          )
        LIMIT 1
        """,
        (LIST_TYPE_CODE_PRINT_RUN_STATUS, raw, raw, raw),
    )
    row = cur.fetchone()
    if row:
        return int(row[0]), None
    return None, "status_not_found"


def main() -> int:
    ap = argparse.ArgumentParser(description="Import print runs by product ISBN.")
    ap.add_argument("--dry-run", action="store_true", help="Validate only and rollback.")
    ap.add_argument(
        "--file",
        "-f",
        default=None,
        help="Override input path (default: PRINT_RUN_INPUT_PATH in this script).",
    )
    args = ap.parse_args()

    path = _resolve_input_path(args.file)
    rows = _read_rows(path)
    if not rows:
        print("No rows found.")
        return 0

    created_by = int(os.environ.get("CREATED_BY_ID", str(CREATED_BY_ID)))
    updated_by_raw = os.environ.get("UPDATED_BY_ID", "")
    updated_by = int(updated_by_raw) if updated_by_raw.strip() else UPDATED_BY_ID

    cnx = mysql.connector.connect(autocommit=False, **_connect_kwargs())
    cur = cnx.cursor()

    inserted = 0
    skipped = 0
    reasons: dict[str, int] = {}
    skipped_rows: list[dict[str, Any]] = []

    def skip(
        reason: str,
        row_no: int,
        isbn: str,
        edition: Any,
        *,
        sheet_title: str = "",
    ) -> None:
        nonlocal skipped
        skipped += 1
        reasons[reason] = reasons.get(reason, 0) + 1
        skipped_rows.append(
            {
                "row_no": row_no,
                "reason": reason,
                "product_isbn": isbn or "",
                "edition_number": str(edition) if edition is not None else "",
                "product_title_ar": sheet_title,
            }
        )

    insert_sql = """
        INSERT INTO inventory_printrun (
          created_at, updated_at,
          product_id, edition_number, price_omr, price, status_id, published_at, notes,
          created_by_id, updated_by_id
        ) VALUES (
          SYSDATE(6), SYSDATE(6),
          %s, %s, %s, %s, %s, %s, %s,
          %s, %s
        )
    """

    try:
        for i, rec in enumerate(rows, start=2):
            isbn_raw = _norm(rec.get("product_isbn")) or ""
            isbn_key = _norm_isbn_key(rec.get("product_isbn"))
            assert isbn_key
            title_hint = _norm(rec.get("product_title_ar")) or ""

            product_id = _resolve_product_id(cur, isbn_key)
            if product_id is None:
                skip(
                    "product_not_found_by_isbn",
                    i,
                    isbn_raw,
                    rec.get("edition_number"),
                    sheet_title=title_hint,
                )
                continue

            edition = _to_int(rec.get("edition_number"))
            if edition is None or edition < 1:
                skip(
                    "invalid_edition_number",
                    i,
                    isbn_raw,
                    rec.get("edition_number"),
                    sheet_title=title_hint,
                )
                continue

            price = _to_decimal(rec.get("price"))
            price_omr = _to_decimal(rec.get("price_omr"))
            if price is None:
                skip("invalid_price", i, isbn_raw, edition, sheet_title=title_hint)
                continue
            if price_omr is None:
                skip("invalid_price_omr", i, isbn_raw, edition, sheet_title=title_hint)
                continue

            pub = _to_date(rec.get("published_at"))
            if pub is None:
                skip(
                    "invalid_or_missing_published_at",
                    i,
                    isbn_raw,
                    edition,
                    sheet_title=title_hint,
                )
                continue

            status_id, st_err = _resolve_status_id(cur, rec.get("status_text"))
            if st_err:
                skip(st_err, i, isbn_raw, edition, sheet_title=title_hint)
                continue

            notes = _norm(rec.get("notes")) or ""

            cur.execute(
                "SELECT id FROM inventory_printrun WHERE product_id = %s AND edition_number = %s LIMIT 1",
                (product_id, edition),
            )
            if cur.fetchone():
                skip(
                    "duplicate_product_edition",
                    i,
                    isbn_raw,
                    edition,
                    sheet_title=title_hint,
                )
                continue

            cur.execute(
                insert_sql,
                (
                    product_id,
                    edition,
                    price_omr,
                    price,
                    status_id,
                    pub,
                    notes,
                    created_by,
                    updated_by,
                ),
            )
            inserted += 1

        if args.dry_run:
            cnx.rollback()
        else:
            cnx.commit()

        print(f"{'Dry-run' if args.dry_run else 'Done'}: total={len(rows)}, inserted={inserted}, skipped={skipped}")
        if reasons:
            print("Skipped reasons:")
            for k, v in sorted(reasons.items(), key=lambda x: (-x[1], x[0])):
                print(f"  - {k}: {v}")

        not_found = [r for r in skipped_rows if r["reason"] == "product_not_found_by_isbn"]
        if not_found:
            print(
                f"\nProducts not found by ISBN ({len(not_found)} row(s)); "
                "name from sheet column product_title_ar:"
            )
            for r in not_found:
                t = r.get("product_title_ar") or "(empty)"
                print(f"  - row~{r['row_no']}: isbn={r['product_isbn']!r} → {t}")

        if skipped_rows:
            print("\nSkipped rows details:")
            for r in skipped_rows:
                tit = r.get("product_title_ar") or ""
                tit_part = f", title={tit!r}" if tit else ""
                print(
                    f"  - row~{r['row_no']}: reason={r['reason']}, "
                    f"isbn={r['product_isbn']!r}, edition={r['edition_number']!r}{tit_part}"
                )

    except Exception as e:
        cnx.rollback()
        print(f"Rolled back. Error: {e}", file=sys.stderr)
        return 1
    finally:
        cur.close()
        cnx.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
