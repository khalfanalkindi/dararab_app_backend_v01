#!/usr/bin/env python3
"""
Update inventory_product.cover_design from spreadsheet rows using ISBN.

Expected columns:
  - isbn (or product_isbn)
  - cover_url (or cover / cover_design)

Default input path is COVER_IMPORT_INPUT_PATH below (same folder as this script).
You can override with --file.

Usage:
  python3 scripts/import_product_covers_mysql.py --dry-run
  python3 scripts/import_product_covers_mysql.py
  python3 scripts/import_product_covers_mysql.py --file /path/to/covers.numbers
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence
from urllib.parse import unquote, urlparse

try:
    import mysql.connector
except ImportError as e:
    print("Install: pip install mysql-connector-python", file=sys.stderr)
    raise e


MYSQL_URL = (
    "mysql://root:DPMCwYKeEPBvcGdQheifhQqICRUuGzBL@mainline.proxy.rlwy.net:29174/railway"
)
MYSQL_HOST = "mainline.proxy.rlwy.net"
MYSQL_PORT = 29174
MYSQL_USER = "root"
MYSQL_PASSWORD = "DPMCwYKeEPBvcGdQheifhQqICRUuGzBL"
MYSQL_DATABASE = "railway"
UPDATED_BY_ID: Optional[int] = None

_SCRIPT_DIR = Path(__file__).resolve().parent
COVER_IMPORT_INPUT_PATH = _SCRIPT_DIR / "product_covers_import.numbers"

HEADER_ALIASES = {
    "isbn": "isbn",
    "product_isbn": "isbn",
    "cover_url": "cover_url",
    "cover": "cover_url",
    "cover_design": "cover_url",
}
REQUIRED_CANONICAL_HEADERS = {"isbn", "cover_url"}


def _parse_mysql_url(url: str) -> dict[str, Any]:
    url = (url or "").strip()
    if not url:
        return {}
    if url.startswith("mysql+pymysql://"):
        url = "mysql://" + url[len("mysql+pymysql://") :]
    elif url.startswith("mysql2://"):
        url = "mysql://" + url[len("mysql2://") :]
    elif not url.startswith("mysql://"):
        return {}
    u = urlparse(url)
    if not u.hostname:
        return {}
    db = (u.path or "").lstrip("/").split("?", 1)[0]
    return {
        "host": u.hostname,
        "port": int(u.port or 3306),
        "user": unquote(u.username or ""),
        "password": unquote(u.password or ""),
        "database": db,
    }


def _connect_kwargs() -> dict[str, Any]:
    raw = os.environ.get("MYSQL_URL") or os.environ.get("DATABASE_URL") or MYSQL_URL
    p = _parse_mysql_url(str(raw))
    return {
        "host": os.environ.get("MYSQL_HOST") or p.get("host") or MYSQL_HOST,
        "port": int(os.environ.get("MYSQL_PORT") or p.get("port") or MYSQL_PORT),
        "user": os.environ.get("MYSQL_USER") or p.get("user") or MYSQL_USER,
        "password": os.environ.get("MYSQL_PASSWORD") or p.get("password") or MYSQL_PASSWORD,
        "database": os.environ.get("MYSQL_DATABASE") or p.get("database") or MYSQL_DATABASE,
        "charset": "utf8mb4",
        "use_unicode": True,
    }


def _resolve_input_path(arg_file: Optional[str]) -> Path:
    if arg_file:
        return Path(arg_file).expanduser().resolve()
    if COVER_IMPORT_INPUT_PATH.is_file():
        return COVER_IMPORT_INPUT_PATH.resolve()
    raise SystemExit(
        f"No cover import file found at:\n  {COVER_IMPORT_INPUT_PATH}\n"
        "Copy your file there, edit COVER_IMPORT_INPUT_PATH in script, or pass --file."
    )


def _norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = " ".join(str(v).split()).strip()
    return s or None


def _norm_isbn_key(v: Any) -> Optional[str]:
    s = _norm(v)
    if not s:
        return None
    return "".join(c for c in s.lower() if c.isalnum())


def _is_valid_url(v: str) -> bool:
    try:
        p = urlparse(v.strip())
    except Exception:
        return False
    return p.scheme in ("http", "https") and bool(p.netloc)


def _canonical_headers(header_row: Sequence[Any]) -> list[str]:
    headers: list[str] = []
    for h in header_row:
        key = "" if h is None else str(h).strip().lower()
        headers.append(HEADER_ALIASES.get(key, key))
    return headers


def _matrix_to_rows(header_row: Sequence[Any], body_rows: Iterable[Sequence[Any]]) -> list[dict[str, Any]]:
    headers = _canonical_headers(header_row)
    missing = REQUIRED_CANONICAL_HEADERS - set(headers)
    if missing:
        raise ValueError(f"Missing required headers: {sorted(missing)}. Found: {headers}")
    idx = {h: i for i, h in enumerate(headers) if h}
    out: list[dict[str, Any]] = []
    for r in body_rows:
        row = list(r or [])
        if not row or all(_norm(x) is None for x in row):
            continue
        isbn_i = idx.get("isbn")
        cover_i = idx.get("cover_url")
        rec = {
            "isbn": row[isbn_i] if isbn_i is not None and isbn_i < len(row) else None,
            "cover_url": row[cover_i] if cover_i is not None and cover_i < len(row) else None,
        }
        if _norm(rec["isbn"]) is None and _norm(rec["cover_url"]) is None:
            continue
        out.append(rec)
    return out


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("CSV has no header row")
            headers = [h for h in reader.fieldnames if h]
            matrix = [[row.get(h) for h in headers] for row in reader]
            return _matrix_to_rows(headers, matrix)
    if suf in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise SystemExit("For .xlsx install: pip install openpyxl") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            wb.close()
            raise ValueError("Empty workbook")
        data = _matrix_to_rows(header, list(rows))
        wb.close()
        return data
    if suf == ".numbers":
        try:
            from numbers_parser import Document
        except ImportError as e:
            raise SystemExit("For .numbers install: pip install numbers-parser") from e
        doc = Document(str(path))
        if not doc.sheets or not doc.sheets[0].tables:
            raise ValueError("Numbers file needs at least one sheet/table")
        matrix = doc.sheets[0].tables[0].rows(values_only=True)
        if not matrix:
            raise ValueError("Numbers table is empty")
        return _matrix_to_rows(matrix[0], matrix[1:])
    raise ValueError(f"Unsupported input file type: {path.suffix!r}")


def main() -> int:
    ap = argparse.ArgumentParser(description="Update product cover_design by ISBN.")
    ap.add_argument("--dry-run", action="store_true", help="Validate only and rollback.")
    ap.add_argument(
        "--file",
        "-f",
        default=None,
        help="Override input path (default: COVER_IMPORT_INPUT_PATH in this script).",
    )
    args = ap.parse_args()

    path = _resolve_input_path(args.file)
    rows = _read_rows(path)
    if not rows:
        print("No rows found.")
        return 0

    updated_by_raw = os.environ.get("UPDATED_BY_ID", "").strip()
    updated_by = int(updated_by_raw) if updated_by_raw else UPDATED_BY_ID

    cnx = mysql.connector.connect(autocommit=False, **_connect_kwargs())
    cur = cnx.cursor()

    updated = 0
    unchanged = 0
    skipped = 0
    reasons: dict[str, int] = {}
    skipped_rows: list[dict[str, Any]] = []

    def skip(reason: str, row_no: int, isbn: str, cover_url: str) -> None:
        nonlocal skipped
        skipped += 1
        reasons[reason] = reasons.get(reason, 0) + 1
        skipped_rows.append(
            {
                "row_no": row_no,
                "reason": reason,
                "isbn": isbn,
                "cover_url": cover_url,
            }
        )

    try:
        for i, rec in enumerate(rows, start=2):
            isbn_raw = _norm(rec.get("isbn")) or ""
            cover_url = _norm(rec.get("cover_url")) or ""
            isbn_key = _norm_isbn_key(isbn_raw)

            if not isbn_key:
                skip("missing_isbn", i, isbn_raw, cover_url)
                continue
            if not cover_url:
                skip("missing_cover_url", i, isbn_raw, cover_url)
                continue
            if not _is_valid_url(cover_url):
                skip("invalid_cover_url", i, isbn_raw, cover_url)
                continue

            cur.execute(
                """
                SELECT id, cover_design
                FROM inventory_product
                WHERE REPLACE(REPLACE(LOWER(TRIM(isbn)), '-', ''), ' ', '') = %s
                LIMIT 1
                """,
                (isbn_key,),
            )
            row = cur.fetchone()
            if not row:
                skip("product_not_found_by_isbn", i, isbn_raw, cover_url)
                continue

            product_id = int(row[0])
            existing_cover = _norm(row[1]) or ""
            if existing_cover == cover_url:
                unchanged += 1
                continue

            cur.execute(
                """
                UPDATE inventory_product
                SET cover_design = %s, updated_at = SYSDATE(6), updated_by_id = %s
                WHERE id = %s
                """,
                (cover_url, updated_by, product_id),
            )
            updated += 1

        if args.dry_run:
            cnx.rollback()
        else:
            cnx.commit()

        print(
            f"{'Dry-run' if args.dry_run else 'Done'}: total={len(rows)}, "
            f"updated={updated}, unchanged={unchanged}, skipped={skipped}"
        )
        if reasons:
            print("Skipped reasons:")
            for k, v in sorted(reasons.items(), key=lambda x: (-x[1], x[0])):
                print(f"  - {k}: {v}")
        if skipped_rows:
            print("\nSkipped rows details:")
            for r in skipped_rows:
                print(
                    f"  - row~{r['row_no']}: reason={r['reason']}, "
                    f"isbn={r['isbn']!r}, cover_url={r['cover_url']!r}"
                )

    except Exception as e:
        cnx.rollback()
        print(f"Rolled back. Error: {e}", file=sys.stderr)
        return 1
    finally:
        cur.close()
        cnx.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
