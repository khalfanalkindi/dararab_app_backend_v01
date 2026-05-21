#!/usr/bin/env python3
"""
One-off: import project rows from a spreadsheet file into MySQL (inventory_project).

**Numbers (Mac):** save as **`projects_import.numbers`** (or **`projects_import.csv`**)
  next to this script under `scripts/`, **or** in the directory you run Python from (repo root is typical).
  Auto-detect checks **run folder first**, then **`scripts/`**. Direct `.numbers` read needs:
  `pip install numbers-parser` (see numbers-parser README for snappy on Mac).

**.xlsx:** optional; requires `pip install openpyxl` (lazy-imported only for .xlsx).

**Single transaction:** all rows commit together, or full rollback on any error.

Usage (typical):
  # 1) Once: create template (+ optional reference CSVs from DB)
  python scripts/import_projects_excel_mysql.py --export-template
  python scripts/import_projects_excel_mysql.py --export-reference   # needs MYSQL_*

  # 2) Align sheet text with DB (recommended once):
  python scripts/import_projects_excel_mysql.py --audit-names
  # 3) Import:
  python scripts/import_projects_excel_mysql.py --dry-run
  python scripts/import_projects_excel_mysql.py

  Connection: see MYSQL_URL / MYSQL_* in the script. Optional shell overrides:
  MYSQL_URL, DATABASE_URL, MYSQL_HOST, MYSQL_PORT, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DATABASE,
  PROJECT_IMPORT_FILE / EXCEL_PATH, CREATED_BY_ID, UPDATED_BY_ID

Spreadsheet columns: see HEADER_ORDER below. **reviewer_name** is optional (column may be omitted).
Project **language_text** resolves against common_listtype.code = **projects_language** (see PROJECT_LANGUAGE_LIST_TYPE_CODE).
"""

from __future__ import annotations

import argparse
import csv
import difflib
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence
from urllib.parse import unquote, urlparse

try:
    import mysql.connector
except ImportError as e:
    print("Install: pip install mysql-connector-python", file=sys.stderr)
    raise e

# =============================================================================
# One-off MySQL / audit defaults — edit for your run.
# Priority: non-empty env (MYSQL_* or MYSQL_URL / DATABASE_URL) > MYSQL_URL below > discrete MYSQL_*.
# Do not commit real passwords to a public repo; use env on CI and rotate if leaked.
# =============================================================================
# Railway / single URL form (leave "" to use MYSQL_HOST / MYSQL_USER / … only):
MYSQL_URL = (
    "mysql://root:DPMCwYKeEPBvcGdQheifhQqICRUuGzBL@mainline.proxy.rlwy.net:29174/railway"
)
MYSQL_HOST = "mainline.proxy.rlwy.net"
MYSQL_PORT = 29174
MYSQL_USER = "root"
MYSQL_PASSWORD = "DPMCwYKeEPBvcGdQheifhQqICRUuGzBL"
MYSQL_DATABASE = "railway"
CREATED_BY_ID = 4
UPDATED_BY_ID = None  # e.g. 4 if your DB disallows NULL on updated_by_id

# Default fallback parties for projects when sheet names are omitted or unmatched.
DEFAULT_AUTHOR_ID = 50
DEFAULT_TRANSLATOR_ID = 15
DEFAULT_RIGHTS_OWNER_ID = 2


def _parse_mysql_url(url: str) -> dict[str, Any]:
    """Parse mysql://user:pass@host:port/dbname into connection fields."""
    url = (url or "").strip()
    if not url:
        return {}
    if url.startswith("mysql+pymysql://"):
        url = "mysql://" + url[len("mysql+pymysql://") :]
    elif url.startswith("mysql2://"):
        url = "mysql://" + url[len("mysql2://") :]
    elif not url.startswith("mysql://"):
        return {}
    u = urlparse(url)
    if not u.hostname:
        return {}
    path = (u.path or "").lstrip("/")
    database = path.split("?", 1)[0] if path else ""
    return {
        "host": u.hostname,
        "port": int(u.port or 3306),
        "user": unquote(u.username or ""),
        "password": unquote(u.password or ""),
        "database": database,
    }


def _url_defaults() -> dict[str, Any]:
    raw = os.environ.get("MYSQL_URL") or os.environ.get("DATABASE_URL") or MYSQL_URL
    return _parse_mysql_url(str(raw))


def _chain_str(env_key: str, url_val: Optional[str], code_fallback: str) -> str:
    v = os.environ.get(env_key)
    if v is not None and str(v).strip() != "":
        return str(v).strip()
    if url_val is not None and str(url_val).strip() != "":
        return str(url_val).strip()
    return code_fallback


def _chain_int(env_key: str, url_val: Optional[int], code_fallback: int) -> int:
    v = os.environ.get(env_key)
    if v is not None and str(v).strip() != "":
        return int(v.strip())
    if url_val is not None:
        return int(url_val)
    return code_fallback


def _cfg_int(env_key: str, code_fallback: int) -> int:
    """Env-only int (no URL part)."""
    return _chain_int(env_key, None, code_fallback)


def _cfg_optional_int(env_key: str, code_fallback: Optional[int]) -> Optional[int]:
    v = os.environ.get(env_key)
    if v is not None and str(v).strip() != "":
        return int(v.strip())
    return code_fallback


def _mysql_connect_kwargs() -> dict[str, Any]:
    p = _url_defaults()
    return {
        "host": _chain_str("MYSQL_HOST", p.get("host"), MYSQL_HOST),
        "port": _chain_int("MYSQL_PORT", p.get("port"), MYSQL_PORT),
        "user": _chain_str("MYSQL_USER", p.get("user"), MYSQL_USER),
        "password": _chain_str("MYSQL_PASSWORD", p.get("password"), MYSQL_PASSWORD),
        "database": _chain_str("MYSQL_DATABASE", p.get("database"), MYSQL_DATABASE),
        "charset": "utf8mb4",
        "use_unicode": True,
    }


def _created_by_id() -> int:
    return _cfg_int("CREATED_BY_ID", CREATED_BY_ID)


def _updated_by_id() -> Optional[int]:
    return _cfg_optional_int("UPDATED_BY_ID", UPDATED_BY_ID)


# Single source of truth for import columns (order = template / CSV columns).
HEADER_ORDER = [
    "title_ar",
    "title_original",
    "manuscript",
    "description",
    "approval_status",
    "progress_status_text",
    "status_text",
    "type_text",
    "language_text",
    "author_name",
    "translator_name",
    "rights_owner_name",
    "reviewer_name",
]

REQUIRED_HEADERS = frozenset(HEADER_ORDER)
# Columns that must appear as headers in the sheet (others may be omitted).
OPTIONAL_SHEET_HEADERS = frozenset(
    {"author_name", "translator_name", "rights_owner_name", "reviewer_name"}
)
MANDATORY_SHEET_HEADERS = REQUIRED_HEADERS - OPTIONAL_SHEET_HEADERS

# common_listtype.code for rows linked from inventory_project.language_id (your DB: projects_language).
PROJECT_LANGUAGE_LIST_TYPE_CODE = "projects_language"

SAMPLE_ROW = {
    "title_ar": "Abu L-’Abbas’s Neighbors",
    "title_original": "جيران أبي العباس",
    "manuscript": "",
    "description": "",
    "approval_status": "TRUE",
    "progress_status_text": "Completed",
    "status_text": "Finalized",
    "type_text": "From Arabic",
    "language_text": "English",
    "author_name": "Ahmad Toufiq",
    "translator_name": "Roger Allen",
    "rights_owner_name": "",
    "reviewer_name": "",
}


_SCRIPT_DIR = Path(__file__).resolve().parent


def _default_reference_dir() -> Path:
    return _SCRIPT_DIR / "projects_import_reference"


def _norm_cell(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    if isinstance(v, bool):
        return str(v)
    return str(v).strip() or None


def _parse_bool(v: Any) -> bool:
    if v is None or v == "":
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "t")


def _list_item_id(
    cur, list_type_code: str, label: Optional[str], *, allow_relaxed: bool = True
) -> Optional[int]:
    if not label:
        return None
    label = label.strip()

    def _run(where_sql: str, params: tuple) -> list:
        cur.execute(
            f"""
            SELECT li.id
            FROM common_listitem li
            INNER JOIN common_listtype lt ON lt.id = li.list_type_id AND lt.code = %s
            WHERE {where_sql}
            LIMIT 2
            """,
            params,
        )
        return cur.fetchall()

    rows = _run(
        "li.display_name_en <=> %s OR li.display_name_ar <=> %s OR li.value <=> %s",
        (list_type_code, label, label, label),
    )
    if len(rows) > 1:
        raise ValueError(
            f"Ambiguous list item for type={list_type_code!r} label={label!r}: {len(rows)} matches"
        )
    if rows:
        return int(rows[0][0])
    if not allow_relaxed:
        return None
    # Case / whitespace tolerant (e.g. "english" vs "English", stray spaces)
    rows = _run(
        """
        LOWER(TRIM(li.display_name_en)) = LOWER(TRIM(%s))
        OR LOWER(TRIM(COALESCE(li.display_name_ar, ''))) = LOWER(TRIM(%s))
        OR LOWER(TRIM(li.value)) = LOWER(TRIM(%s))
        """,
        (list_type_code, label, label, label),
    )
    if len(rows) > 1:
        raise ValueError(
            f"Ambiguous list item for type={list_type_code!r} label={label!r}: {len(rows)} matches"
        )
    if rows:
        return int(rows[0][0])
    return None


def _list_item_canonical_label(cur, list_item_id: int) -> str:
    cur.execute(
        "SELECT COALESCE(NULLIF(TRIM(display_name_en), ''), NULLIF(TRIM(value), '')) "
        "FROM common_listitem WHERE id = %s",
        (list_item_id,),
    )
    r = cur.fetchone()
    return str(r[0]) if r and r[0] is not None else ""


_TABLES = frozenset(
    {"inventory_author", "inventory_translator", "inventory_rightsowner", "inventory_reviewer"}
)


def _party_name_variants(name: str) -> list[str]:
    """Try spreadsheet variants against DB names (spacing, common typos)."""
    n = " ".join(str(name).split()).strip()
    out: list[str] = []
    seen: set[str] = set()

    def add(s: str) -> None:
        if s and s not in seen:
            seen.add(s)
            out.append(s)

    add(n)
    add(n.replace("Christiaan", "Christian"))
    add(n.replace("Al Muqri", "Al-Muqri"))
    add(n.replace("Al-Muqri", "Al Muqri"))
    add(n.replace("Ahmed", "Ahmad"))
    add(n.replace("Hussien", "Hussain"))
    add(n.replace("Badryia", "Badria"))
    add(n.replace("Al Shahi", "Alshihi"))
    add(n.replace("Al Shahi", "Al Shihi"))
    add(n.replace("Badryia", "Badria").replace("Al Shahi", "Alshihi"))
    return out


# id -> full name rows, refreshed once per import run
_PARTY_NAME_CACHE: dict[str, list[tuple[int, str]]] = {}


def _party_table_pairs(cur, table: str) -> list[tuple[int, str]]:
    if table not in _PARTY_NAME_CACHE:
        cur.execute(f"SELECT id, name FROM `{table}` ORDER BY id")
        _PARTY_NAME_CACHE[table] = [(int(r[0]), str(r[1])) for r in cur.fetchall()]
    return _PARTY_NAME_CACHE[table]


def _party_id_fuzzy(cur, table: str, name: str) -> Optional[int]:
    """Match sheet name to closest DB name when exact variants fail (transliteration drift)."""
    pairs = _party_table_pairs(cur, table)
    if not pairs:
        return None
    db_names = [p[1] for p in pairs]
    hits = difflib.get_close_matches(name, db_names, n=1, cutoff=0.78)
    if not hits:
        return None
    best = hits[0]
    for pid, dbn in pairs:
        if dbn == best:
            return pid
    return None


def _party_id(cur, table: str, name: Optional[str], *, allow_fuzzy: bool = True) -> Optional[int]:
    if not name:
        return None
    if table not in _TABLES:
        raise ValueError(f"Invalid party table: {table}")
    name = name.strip()
    for candidate in _party_name_variants(name):
        cur.execute(f"SELECT id FROM `{table}` WHERE name = %s LIMIT 2", (candidate,))
        rows = cur.fetchall()
        if len(rows) > 1:
            raise ValueError(f"Ambiguous {table} name={name!r} (matched {candidate!r})")
        if rows:
            return int(rows[0][0])
    if not allow_fuzzy:
        return None
    fid = _party_id_fuzzy(cur, table, name)
    if fid is not None:
        return fid
    return None


def _party_db_name_for_id(cur, table: str, row_id: int) -> str:
    cur.execute(f"SELECT name FROM `{table}` WHERE id = %s", (row_id,))
    r = cur.fetchone()
    return str(r[0]) if r and r[0] is not None else ""


def _matrix_to_records(header_row: Sequence[Any], body_rows: Iterable[Sequence[Any]]) -> list[dict[str, Any]]:
    """First row = headers; following rows = data. Optional columns may be absent (see OPTIONAL_SHEET_HEADERS)."""
    headers = ["" if h is None else str(h).strip().lower() for h in header_row]
    missing = MANDATORY_SHEET_HEADERS - set(headers)
    if missing:
        raise ValueError(f"Table missing columns: {sorted(missing)}. Found headers: {headers}")

    col_index = {h: i for i, h in enumerate(headers) if h}
    col_known = {k: col_index[k] for k in HEADER_ORDER if k in col_index}
    out: list[dict[str, Any]] = []
    for excel_row in body_rows:
        if excel_row is None:
            continue
        row_list = list(excel_row)
        if not row_list or all(v is None or str(v).strip() == "" for v in row_list):
            continue
        rec: dict[str, Any] = {}
        for key in HEADER_ORDER:
            idx = col_known.get(key)
            if idx is None:
                rec[key] = None
            else:
                rec[key] = row_list[idx] if idx < len(row_list) else None
        if not _norm_cell(rec.get("title_ar")):
            continue
        out.append(rec)
    return out


def _read_rows_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("For .xlsx install: pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("Empty sheet")

    try:
        return _matrix_to_records(header_row, rows_iter)
    finally:
        wb.close()


def _read_rows_numbers(path: Path) -> list[dict[str, Any]]:
    try:
        from numbers_parser import Document
    except ImportError as e:
        raise SystemExit(
            "For .numbers install: pip install numbers-parser\n"
            "On Mac you may need: brew install snappy (see https://github.com/masaccio/numbers-parser)"
        ) from e

    doc = Document(str(path))
    if not doc.sheets:
        raise ValueError("Numbers document has no sheets")
    sheet = doc.sheets[0]
    if not sheet.tables:
        raise ValueError(f"Numbers sheet {sheet.name!r} has no tables — add a table with the header row.")
    table = sheet.tables[0]
    matrix = table.rows(values_only=True)
    if not matrix:
        raise ValueError("Numbers table is empty")
    header_row = matrix[0]
    body_rows = matrix[1:]
    return _matrix_to_records(header_row, body_rows)


def _read_rows_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("CSV has no header row")
        lower_map: dict[str, str] = {}
        for fn in reader.fieldnames:
            if fn is None:
                continue
            raw = fn.strip()
            lo = raw.lower()
            if lo in REQUIRED_HEADERS:
                lower_map[lo] = raw
        missing = MANDATORY_SHEET_HEADERS - set(lower_map.keys())
        if missing:
            raise ValueError(f"CSV missing columns: {sorted(missing)}. Found: {reader.fieldnames}")

        out: list[dict[str, Any]] = []
        for row in reader:
            rec = {
                k: (row.get(lower_map[k]) if k in lower_map else None) for k in HEADER_ORDER
            }
            if not _norm_cell(rec.get("title_ar")):
                continue
            out.append(rec)
    return out


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        return _read_rows_csv(path)
    if suf == ".numbers":
        return _read_rows_numbers(path)
    if suf in (".xlsx", ".xlsm"):
        return _read_rows_xlsx(path)
    raise ValueError(
        f"Unsupported file type {path.suffix!r}; use .numbers, .csv, or .xlsx"
    )


def cmd_export_template(dest: Path, include_sample: bool) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=HEADER_ORDER, extrasaction="ignore")
        w.writeheader()
        if include_sample:
            w.writerow({k: SAMPLE_ROW[k] for k in HEADER_ORDER})
    print(f"Wrote template: {dest.resolve()}")
    print(
        "Open in Numbers → fill rows → save as projects_import.numbers (or export CSV UTF-8) "
        "in the folder you run from, or under scripts/ next to this file. "
        "Optional: pip install numbers-parser to read .numbers directly."
    )


def cmd_export_reference(out_dir: Path) -> None:
    kw = _mysql_connect_kwargs()
    user = kw["user"]
    database = kw["database"]
    if not user or not database:
        raise SystemExit(
            "Fill MYSQL_USER / MYSQL_DATABASE (and password) in the script config block, "
            "or set them in the environment, for --export-reference."
        )

    out_dir.mkdir(parents=True, exist_ok=True)
    cnx = mysql.connector.connect(**kw)
    cur = cnx.cursor()

    def dump_list_type(code: str, fname: str) -> None:
        cur.execute(
            """
            SELECT li.id, li.value, li.display_name_en, li.display_name_ar
            FROM common_listitem li
            JOIN common_listtype lt ON lt.id = li.list_type_id AND lt.code = %s
            ORDER BY li.id
            """,
            (code,),
        )
        rows = cur.fetchall()
        p = out_dir / fname
        with p.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["id", "value", "display_name_en", "display_name_ar"])
            w.writerows(rows)
        print(f"Wrote {len(rows)} rows → {p}")

    def dump_table(table: str, fname: str) -> None:
        if table not in _TABLES:
            raise ValueError(table)
        cur.execute(f"SELECT id, name FROM `{table}` ORDER BY name")
        rows = cur.fetchall()
        p = out_dir / fname
        with p.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["id", "name"])
            w.writerows(rows)
        print(f"Wrote {len(rows)} rows → {p}")

    try:
        dump_list_type("progress_status", "list_progress_status.csv")
        dump_list_type("projects_status", "list_projects_status.csv")
        dump_list_type("projects_type", "list_projects_type.csv")
        dump_list_type(PROJECT_LANGUAGE_LIST_TYPE_CODE, "list_projects_language.csv")
        dump_list_type("language", "list_language.csv")
        dump_table("inventory_author", "authors.csv")
        dump_table("inventory_translator", "translators.csv")
        dump_table("inventory_rightsowner", "rights_owners.csv")
        dump_table("inventory_reviewer", "reviewers.csv")
    finally:
        cur.close()
        cnx.close()

    print(f"Reference CSVs in: {out_dir.resolve()}")


def _lov_sample_labels(cur, list_type_code: str, limit: int = 14) -> str:
    cur.execute(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(li.display_name_en), ''), TRIM(li.value))
        FROM common_listitem li
        INNER JOIN common_listtype lt ON lt.id = li.list_type_id AND lt.code = %s
        WHERE li.is_active = 1
        ORDER BY 1
        LIMIT %s
        """,
        (list_type_code, limit),
    )
    vals = [str(r[0]) for r in cur.fetchall() if r[0]]
    return ", ".join(vals) if vals else "(no active list items)"


def cmd_audit_names(args: argparse.Namespace) -> int:
    """
    Print how sheet values compare to DB (strict match vs fuzzy/relaxed).
    Use this to fix Numbers/CSV text to exactly match the system, then import without guesswork.
    """
    path = _resolve_import_path(args.file)
    if not path.is_file():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    kw = _mysql_connect_kwargs()
    if not kw["user"] or not kw["database"]:
        print("Configure MySQL (MYSQL_URL or MYSQL_* in the script).", file=sys.stderr)
        return 1

    rows = _read_rows(path)
    if not rows:
        print("No data rows in import file.")
        return 0

    cnx = mysql.connector.connect(**kw)
    cur = cnx.cursor()
    _PARTY_NAME_CACHE.clear()

    print(f"Audit: {path.name} ({len(rows)} data rows)\n")

    party_specs: list[tuple[str, str]] = [
        ("author_name", "inventory_author"),
        ("translator_name", "inventory_translator"),
        ("rights_owner_name", "inventory_rightsowner"),
        ("reviewer_name", "inventory_reviewer"),
    ]
    for field, table in party_specs:
        values = sorted({_norm_cell(rec[field]) for rec in rows if _norm_cell(rec.get(field))})
        if not values:
            continue
        print(f"### {field}  →  {table}")
        for v in values:
            assert v is not None
            strict_id = _party_id(cur, table, v, allow_fuzzy=False)
            full_id = _party_id(cur, table, v, allow_fuzzy=True)
            if strict_id:
                canon = _party_db_name_for_id(cur, table, strict_id)
                if canon == v or canon.casefold() == v.casefold():
                    print(f"  OK     {v!r}")
                else:
                    print(f"  FIX    sheet={v!r}  →  use exactly: {canon!r}")
            elif full_id:
                canon = _party_db_name_for_id(cur, table, full_id)
                print(f"  FIX    sheet={v!r}  →  use exactly: {canon!r}  (fuzzy match)")
            else:
                print(f"  MISS   {v!r}  (add in DB or fix spelling)")
                print(f"         sample DB names: {_party_name_examples(cur, table)}")
        print()

    lov_specs: list[tuple[str, str]] = [
        ("progress_status_text", "progress_status"),
        ("status_text", "projects_status"),
        ("type_text", "projects_type"),
        ("language_text", PROJECT_LANGUAGE_LIST_TYPE_CODE),
    ]
    for field, list_code in lov_specs:
        values = sorted({_norm_cell(rec[field]) for rec in rows if _norm_cell(rec.get(field))})
        if not values:
            continue
        print(f"### {field}  →  common_listitem (code={list_code!r})")
        for v in values:
            assert v is not None
            strict_id = _list_item_id(cur, list_code, v, allow_relaxed=False)
            full_id = _list_item_id(cur, list_code, v, allow_relaxed=True)
            if strict_id:
                lab = _list_item_canonical_label(cur, strict_id)
                if lab == v or lab.casefold() == v.casefold():
                    print(f"  OK     {v!r}")
                else:
                    print(f"  FIX    sheet={v!r}  →  use exactly: {lab!r}")
            elif full_id:
                lab = _list_item_canonical_label(cur, full_id)
                print(f"  FIX    sheet={v!r}  →  use exactly: {lab!r}  (case/trim match)")
            else:
                print(f"  MISS   {v!r}")
                print(f"         try one of: {_lov_sample_labels(cur, list_code)}")
        print()

    cur.close()
    cnx.close()
    print('Update your sheet so every FIX line matches the "use exactly" value, then run --dry-run.')
    return 0


def _party_name_examples(cur, table: str) -> str:
    cur.execute(f"SELECT name FROM `{table}` ORDER BY name LIMIT 12")
    return ", ".join(str(r[0]) for r in cur.fetchall())


def _resolve_import_path(arg_file: Optional[str]) -> Path:
    env_path = os.environ.get("PROJECT_IMPORT_FILE") or os.environ.get("EXCEL_PATH")
    if arg_file:
        return Path(arg_file).expanduser().resolve()
    if env_path:
        return Path(env_path).expanduser().resolve()
    # Default names: prefer .numbers, then .csv. Search cwd first (good for repo root + .gitignore), then scripts/.
    for fname in ("projects_import.numbers", "projects_import.csv"):
        for base in (Path.cwd(), _SCRIPT_DIR):
            p = base / fname
            if p.is_file():
                return p.resolve()
    raise SystemExit(
        "No import file. Use one of:\n"
        "  --file /path/to/projects_import.numbers   (or .csv / .xlsx)\n"
        "  export PROJECT_IMPORT_FILE=...   (or EXCEL_PATH)\n"
        "  Or place projects_import.numbers / projects_import.csv in the current directory or in scripts/\n"
        "  (.numbers wins over .csv if both exist in the same folder; cwd is checked before scripts/).\n"
        "Create a template: python scripts/import_projects_excel_mysql.py --export-template"
    )


def cmd_import(args: argparse.Namespace) -> int:
    path = _resolve_import_path(args.file)
    if not path.is_file():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    kw = _mysql_connect_kwargs()
    user = kw["user"]
    database = kw["database"]
    created_by = _created_by_id()
    updated_by = _updated_by_id()

    if not user or not database:
        print(
            "Fill MYSQL_USER / MYSQL_DATABASE (and password) in the script config block, "
            "or set them in the environment.",
            file=sys.stderr,
        )
        return 1

    rows = _read_rows(path)
    if not rows:
        print("No data rows found (need title_ar on each row).")
        return 0

    cnx = mysql.connector.connect(autocommit=False, **kw)
    cur = cnx.cursor()
    _PARTY_NAME_CACHE.clear()
    now = datetime.now()

    insert_sql = """
        INSERT INTO inventory_project (
            created_at, updated_at,
            title_ar, title_original, manuscript, description, approval_status,
            progress_status_id, status_id, type_id, language_id,
            author_id, translator_id, rights_owner_id, reviewer_id,
            created_by_id, updated_by_id
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
    """

    import_warnings: list[str] = []
    try:
        for i, rec in enumerate(rows, start=2):
            title_ar = _norm_cell(rec["title_ar"])
            title_original = _norm_cell(rec["title_original"]) or ""
            manuscript = _norm_cell(rec["manuscript"]) or ""
            description = _norm_cell(rec["description"]) or ""
            approval = _parse_bool(rec["approval_status"])

            ps = _list_item_id(cur, "progress_status", _norm_cell(rec["progress_status_text"]))
            st = _list_item_id(cur, "projects_status", _norm_cell(rec["status_text"]))
            ty = _list_item_id(cur, "projects_type", _norm_cell(rec["type_text"]))
            lg = _list_item_id(
                cur, PROJECT_LANGUAGE_LIST_TYPE_CODE, _norm_cell(rec["language_text"])
            )

            au_name = _norm_cell(rec["author_name"])
            tr_name = _norm_cell(rec["translator_name"])
            ro_name = _norm_cell(rec["rights_owner_name"])

            au_found = _party_id(cur, "inventory_author", au_name)
            tr_found = _party_id(cur, "inventory_translator", tr_name)
            ro_found = _party_id(cur, "inventory_rightsowner", ro_name)

            au = au_found or DEFAULT_AUTHOR_ID
            tr = tr_found or DEFAULT_TRANSLATOR_ID
            ro = ro_found or DEFAULT_RIGHTS_OWNER_ID
            rv = _party_id(cur, "inventory_reviewer", _norm_cell(rec["reviewer_name"]))

            missing = []
            if _norm_cell(rec["progress_status_text"]) and ps is None:
                missing.append(f"progress_status_text={rec['progress_status_text']!r}")
            if _norm_cell(rec["status_text"]) and st is None:
                missing.append(f"status_text={rec['status_text']!r}")
            if _norm_cell(rec["type_text"]) and ty is None:
                missing.append(f"type_text={rec['type_text']!r}")
            if _norm_cell(rec["language_text"]) and lg is None:
                missing.append(f"language_text={rec['language_text']!r}")
            if au_name and au_found is None:
                import_warnings.append(
                    f"Row ~{i} ({title_ar!r}): author_name={au_name!r} not found — "
                    f"using default author_id={DEFAULT_AUTHOR_ID}."
                )
            if tr_name and tr_found is None:
                import_warnings.append(
                    f"Row ~{i} ({title_ar!r}): translator_name={tr_name!r} not found — "
                    f"using default translator_id={DEFAULT_TRANSLATOR_ID}."
                )
            if ro_name and ro_found is None:
                import_warnings.append(
                    f"Row ~{i} ({title_ar!r}): rights_owner_name={ro_name!r} not found — "
                    f"using default rights_owner_id={DEFAULT_RIGHTS_OWNER_ID}."
                )
            if _norm_cell(rec["reviewer_name"]) and rv is None:
                import_warnings.append(
                    f"Row ~{i} ({title_ar!r}): reviewer_name={_norm_cell(rec['reviewer_name'])!r} "
                    "not in inventory_reviewer — using NULL."
                )

            if missing:
                raise ValueError(f"Row ~{i} ({title_ar!r}): unresolved " + "; ".join(missing))

            params = (
                now,
                now,
                title_ar,
                title_original,
                manuscript,
                description,
                1 if approval else 0,
                ps,
                st,
                ty,
                lg,
                au,
                tr,
                ro,
                rv,
                created_by,
                updated_by,
            )
            if not args.dry_run:
                cur.execute(insert_sql, params)

        if args.dry_run:
            cnx.rollback()
            print(f"Dry-run OK: {len(rows)} rows validated (rolled back, no INSERT).")
        else:
            cnx.commit()
            print(f"Committed {len(rows)} project rows.")
        if import_warnings:
            print("\nWarnings (import still proceeds for these rows):", file=sys.stderr)
            for w in import_warnings:
                print(f"  {w}", file=sys.stderr)
    except Exception as e:
        cnx.rollback()
        print(f"Rolled back. Error: {e}", file=sys.stderr)
        return 1
    finally:
        cur.close()
        cnx.close()

    return 0


def main() -> int:
    p = argparse.ArgumentParser(
        description="Import projects from CSV (Numbers) or .xlsx into MySQL — single transaction."
    )
    p.add_argument("--dry-run", action="store_true", help="Validate only; no INSERT (still connects to DB).")
    p.add_argument(
        "--file",
        "-f",
        default=None,
        metavar="PATH",
        help="Import file .numbers, .csv, or .xlsx (else env path / ./projects_import.numbers / .csv).",
    )
    p.add_argument(
        "--export-template",
        nargs="?",
        const="",
        metavar="PATH",
        help="Write CSV template (optional path; default: scripts/projects_import_template.csv).",
    )
    p.add_argument(
        "--no-sample-row",
        action="store_true",
        help="With --export-template, write header only (no example row).",
    )
    p.add_argument(
        "--export-reference",
        nargs="?",
        const="",
        metavar="DIR",
        help="Dump LOV + party name CSVs for Numbers (optional dir; default: scripts/projects_import_reference/).",
    )
    p.add_argument(
        "--audit-names",
        action="store_true",
        help="Compare sheet LOV/party strings to DB; print FIX lines to edit Numbers/CSV to match the system.",
    )
    args = p.parse_args()

    if args.export_template is not None:
        dest = (
            _SCRIPT_DIR / "projects_import_template.csv"
            if args.export_template == ""
            else Path(args.export_template).expanduser()
        )
        cmd_export_template(dest, include_sample=not args.no_sample_row)
        return 0

    if args.export_reference is not None:
        out = _default_reference_dir() if args.export_reference == "" else Path(args.export_reference).expanduser()
        cmd_export_reference(out)
        return 0

    if args.audit_names:
        return cmd_audit_names(args)

    return cmd_import(args)


if __name__ == "__main__":
    sys.exit(main())
