#!/usr/bin/env python3
"""
One-off: insert missing rights owners from a dedicated rights-owner import file.

Reads rights_owner_name values from:
- .numbers (preferred)
- .csv
- .xlsx / .xlsm

Defaults:
- Reads from scripts/rightowiner.numbers by default
- (or pass --file / set RIGHTS_OWNER_IMPORT_FILE)
- Transactional (rollback on error)
- Skips existing names
- Inserts contact_info='' with created_by_id=4

Usage:
  python scripts/insert_rights_owners_mysql.py --dry-run
  python scripts/insert_rights_owners_mysql.py
  python scripts/insert_rights_owners_mysql.py --file scripts/rightowiner.numbers

This script can also update existing inventory_project rows that still have the
fallback rights_owner_id (default 2), using title_ar + rights_owner_name from
the same input file.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence
from urllib.parse import unquote, urlparse

try:
    import mysql.connector
except ImportError as e:
    print("Install: pip install mysql-connector-python", file=sys.stderr)
    raise e


# Connection defaults (env vars override)
MYSQL_URL = (
    "mysql://root:DPMCwYKeEPBvcGdQheifhQqICRUuGzBL@mainline.proxy.rlwy.net:29174/railway"
)
MYSQL_HOST = "mainline.proxy.rlwy.net"
MYSQL_PORT = 29174
MYSQL_USER = "root"
MYSQL_PASSWORD = "DPMCwYKeEPBvcGdQheifhQqICRUuGzBL"
MYSQL_DATABASE = "railway"
CREATED_BY_ID = 4
UPDATED_BY_ID = None  # set to 4 if needed

TARGET_COLUMN = "rights_owner_name"
TITLE_COLUMN = "title_ar"
_SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_FILE = _SCRIPT_DIR / "rightowiner.numbers"


def _parse_mysql_url(url: str) -> dict[str, Any]:
    url = (url or "").strip()
    if not url:
        return {}
    if url.startswith("mysql+pymysql://"):
        url = "mysql://" + url[len("mysql+pymysql://") :]
    elif url.startswith("mysql2://"):
        url = "mysql://" + url[len("mysql2://") :]
    elif not url.startswith("mysql://"):
        return {}
    u = urlparse(url)
    if not u.hostname:
        return {}
    path = (u.path or "").lstrip("/")
    database = path.split("?", 1)[0] if path else ""
    return {
        "host": u.hostname,
        "port": int(u.port or 3306),
        "user": unquote(u.username or ""),
        "password": unquote(u.password or ""),
        "database": database,
    }


def _connect_kwargs() -> dict[str, Any]:
    url = os.environ.get("MYSQL_URL") or os.environ.get("DATABASE_URL") or MYSQL_URL
    p = _parse_mysql_url(str(url))
    return {
        "host": os.environ.get("MYSQL_HOST") or p.get("host") or MYSQL_HOST,
        "port": int(os.environ.get("MYSQL_PORT") or p.get("port") or MYSQL_PORT),
        "user": os.environ.get("MYSQL_USER") or p.get("user") or MYSQL_USER,
        "password": os.environ.get("MYSQL_PASSWORD") or p.get("password") or MYSQL_PASSWORD,
        "database": os.environ.get("MYSQL_DATABASE") or p.get("database") or MYSQL_DATABASE,
        "charset": "utf8mb4",
        "use_unicode": True,
    }


def _resolve_import_path(arg_file: Optional[str]) -> Path:
    env_path = os.environ.get("RIGHTS_OWNER_IMPORT_FILE")
    if arg_file:
        return Path(arg_file).expanduser().resolve()
    if env_path:
        return Path(env_path).expanduser().resolve()
    if DEFAULT_INPUT_FILE.is_file():
        return DEFAULT_INPUT_FILE.resolve()
    raise SystemExit(
        f"No import file found at {DEFAULT_INPUT_FILE}. "
        "Use --file or set RIGHTS_OWNER_IMPORT_FILE."
    )


def _norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = " ".join(str(v).split()).strip()
    return s or None


def _extract_names_from_matrix(header_row: Sequence[Any], body_rows: Iterable[Sequence[Any]]) -> list[str]:
    headers = ["" if h is None else str(h).strip().lower() for h in header_row]
    if TARGET_COLUMN not in headers:
        raise ValueError(f"Column {TARGET_COLUMN!r} not found. Headers: {headers}")
    idx = headers.index(TARGET_COLUMN)
    seen: set[str] = set()
    out: list[str] = []
    for r in body_rows:
        if r is None:
            continue
        row = list(r)
        val = _norm(row[idx] if idx < len(row) else None)
        if not val:
            continue
        if val not in seen:
            seen.add(val)
            out.append(val)
    return out


def _extract_title_owner_pairs_from_matrix(
    header_row: Sequence[Any], body_rows: Iterable[Sequence[Any]]
) -> list[tuple[str, str]]:
    headers = ["" if h is None else str(h).strip().lower() for h in header_row]
    if TITLE_COLUMN not in headers or TARGET_COLUMN not in headers:
        return []
    t_idx = headers.index(TITLE_COLUMN)
    r_idx = headers.index(TARGET_COLUMN)
    seen: set[tuple[str, str]] = set()
    out: list[tuple[str, str]] = []
    for r in body_rows:
        if r is None:
            continue
        row = list(r)
        title = _norm(row[t_idx] if t_idx < len(row) else None)
        owner = _norm(row[r_idx] if r_idx < len(row) else None)
        if not title or not owner:
            continue
        k = (title, owner)
        if k not in seen:
            seen.add(k)
            out.append(k)
    return out


def _read_matrix(path: Path) -> tuple[Sequence[Any], list[Sequence[Any]]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("CSV has no header row")
            headers = list(reader.fieldnames)
            body: list[Sequence[Any]] = []
            for row in reader:
                body.append([row.get(h) for h in headers])
            return headers, body

    if suf in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise SystemExit("For .xlsx install: pip install openpyxl") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            wb.close()
            raise ValueError("Empty workbook")
        try:
            return header, list(rows)
        finally:
            wb.close()

    if suf == ".numbers":
        try:
            from numbers_parser import Document
        except ImportError as e:
            raise SystemExit(
                "For .numbers install: pip install numbers-parser\n"
                "On Mac you may need: brew install snappy"
            ) from e
        doc = Document(str(path))
        if not doc.sheets or not doc.sheets[0].tables:
            raise ValueError("Numbers file must have at least one sheet and one table")
        matrix = doc.sheets[0].tables[0].rows(values_only=True)
        if not matrix:
            raise ValueError("Numbers table is empty")
        return matrix[0], list(matrix[1:])

    raise ValueError(f"Unsupported file type: {path.suffix!r}")


def _read_names(path: Path) -> list[str]:
    header, body = _read_matrix(path)
    return _extract_names_from_matrix(header, body)


def _read_title_owner_pairs(path: Path) -> list[tuple[str, str]]:
    header, body = _read_matrix(path)
    return _extract_title_owner_pairs_from_matrix(header, body)


def main() -> int:
    p = argparse.ArgumentParser(
        description="Insert missing rights owners from file and sync project rights owners by title."
    )
    p.add_argument("--dry-run", action="store_true", help="Validate and show what would insert, then rollback.")
    p.add_argument("--file", "-f", default=None, help="Path to .numbers/.csv/.xlsx import file.")
    args = p.parse_args()

    path = _resolve_import_path(args.file)
    if not path.is_file():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    names = _read_names(path)
    pairs = _read_title_owner_pairs(path)
    if not names:
        print(f"No non-empty {TARGET_COLUMN!r} values found in {path.name}.")
        return 0

    created_by = int(os.environ.get("CREATED_BY_ID", str(CREATED_BY_ID)))
    updated_by_raw = os.environ.get("UPDATED_BY_ID", "")
    updated_by = int(updated_by_raw) if updated_by_raw.strip() else UPDATED_BY_ID

    cnx = mysql.connector.connect(autocommit=False, **_connect_kwargs())
    cur = cnx.cursor()
    try:
        cur.execute("SELECT name FROM inventory_rightsowner")
        existing = {str(r[0]) for r in cur.fetchall() if r and r[0] is not None}
        to_insert = [n for n in names if n not in existing]

        insert_sql = """
            INSERT INTO inventory_rightsowner
              (created_at, updated_at, name, contact_info, created_by_id, updated_by_id)
            VALUES
              (SYSDATE(6), SYSDATE(6), %s, %s, %s, %s)
        """
        payload = [(n, "", created_by, updated_by) for n in to_insert]

        # Force sync from file mapping, even when project.rights_owner_id is already set.
        # Use normalized text matching to avoid misses from casing/spaces.
        update_sql = """
            UPDATE inventory_project p
            JOIN inventory_rightsowner r
              ON LOWER(TRIM(r.name)) = LOWER(TRIM(%s))
            SET p.rights_owner_id = r.id
            WHERE LOWER(TRIM(p.title_ar)) = LOWER(TRIM(%s))
        """

        synced = 0
        if pairs and args.dry_run:
            # Sync projects directly by title_ar + rights_owner_name mapping.
            for title_ar, owner_name in pairs:
                cur.execute(update_sql, (owner_name, title_ar))
                synced += cur.rowcount

        if args.dry_run:
            cnx.rollback()
            print(
                f"Dry-run: would insert {len(to_insert)} rights owners "
                f"(from {len(names)} names in {path.name}; {len(existing)} already existed)."
            )
            if to_insert:
                print("First 20 to insert:", ", ".join(to_insert[:20]))
            else:
                print(f"No new rights owners to insert; all {len(names)} names already exist.")
            if pairs:
                print(
                    f"Dry-run: would sync {synced} project rows using "
                    f"{len(pairs)} title/owner mappings from {path.name}."
                )
        else:
            if to_insert:
                cur.executemany(insert_sql, payload)
            # sync after inserts so newly added rights owners are available
            synced = 0
            if pairs:
                for title_ar, owner_name in pairs:
                    cur.execute(update_sql, (owner_name, title_ar))
                    synced += cur.rowcount
            cnx.commit()
            if to_insert:
                print(
                    f"Inserted {len(to_insert)} rights owners "
                    f"(from {len(names)} names in {path.name}; {len(existing)} already existed)."
                )
            else:
                print(f"No new rights owners inserted; all {len(names)} names already existed.")
            if pairs:
                print(
                    f"Synced {synced} project rows from {path.name} "
                    "(matched by title_ar + rights_owner_name)."
                )
    except Exception as e:
        cnx.rollback()
        print(f"Rolled back. Error: {e}", file=sys.stderr)
        return 1
    finally:
        cur.close()
        cnx.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())

